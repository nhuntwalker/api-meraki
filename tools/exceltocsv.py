import os, sys, csv, glob
from openpyxl import load_workbook

def main(file):

    print("Separating Worksheets from Workbook")
    sheets = []
    workbook = load_workbook(file, read_only=True, data_only=True)
    all_worksheets = workbook.sheetnames
    for worksheet_name in all_worksheets:
        sheets.append(worksheet_name)
        
        print("Extracting data from Worksheets")
        for worksheet_name in workbook:
           #print("Export " + worksheet_name + " ...")
            worksheet = workbook[(worksheet_name)]

            output_csv = open(''.join([worksheet_name, '.csv']), newline="", mode='wt')
            wr = csv.writer(output_csv, quoting=csv.QUOTE_ALL)
            for row in worksheet.iter_rows():
                lrow = []
                for cell in row:
                    lrow.append(cell.value)
                wr.writerow(lrow)
            output_csv.close()
            print("writing data to csv file")

if __name__ == '__main__':
    os.chdir(os.path.join(os.pardir, 'data'))
    for file in glob.glob('*.xlsx'):
        print(file)
        main(file)
